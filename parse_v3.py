#!/usr/bin/env python3
"""
parse_v3.py — Generates app/src/data/grammarExercises.js
Sources: grammar_theory.docx  +  grammar_exercises.xlsx  +  app/src/data/grammar.js

Run from the project root:
    python3 parse_v3.py
"""

import re
import json
from pathlib import Path

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).parent
DOCX_PATH  = ROOT / "grammar_theory.docx"
XLSX_PATH  = ROOT / "grammar_exercises.xlsx"
GRAMMAR_JS = ROOT / "app" / "src" / "data" / "grammar.js"
OUT_JS     = ROOT / "app" / "src" / "data" / "grammarExercises.js"

LEVELS = ["A1", "A2", "B1", "B2", "C1"]

# ── 1. Load canonical topic order from grammar.js ────────────────────────────
def load_canonical_topics(path: Path) -> dict[str, list[str]]:
    """Returns {level: [topic, ...]} from grammar.js."""
    text = path.read_text(encoding="utf-8")
    result = {}
    for level in LEVELS:
        # Bug-fix #1: no lookahead (fails for the last block C1 which ends with \n};)
        m = re.search(rf'{level}:\s*\[(.*?)\],', text, re.DOTALL)
        if not m:
            print(f"  WARNING: could not find {level} block in grammar.js")
            result[level] = []
            continue
        block = m.group(1)
        # Use escaped-quote-aware regex, then unescape \" → "
        raw_topics = re.findall(r'topic:"((?:[^"\\]|\\.)*)"', block)
        topics = [t.replace('\\"', '"').replace('\\\\', '\\') for t in raw_topics]
        result[level] = topics
    return result

# ── 2. Normalisation for fuzzy matching ──────────────────────────────────────
def _norm(s: str) -> str:
    """Strip whitespace, level suffix, trailing parentheticals, then lowercase."""
    s = s.strip()
    # Bug-fix #3: strip bare level tag FIRST, then parentheticals in loop
    s = re.sub(r'\s+[A-C][12]\s*$', '', s).strip()
    while True:
        new_s = re.sub(r'\s*\([^)]*\)\s*$', '', s).strip()
        if new_s == s:
            break
        s = new_s
    return s.lower()

# ── 3. Parse grammar_theory.docx ─────────────────────────────────────────────
def parse_docx(docx_path: Path, canonical: dict) -> dict:
    """Returns {level: {topic_key: theory_string}}."""
    from docx import Document
    doc = Document(docx_path)

    # Build reverse lookups: exact (lowercase, level-suffix stripped) first,
    # then parenthetical-stripped fallback.  Exact wins so that topics like
    # "Personal Pronouns (Nominative)" and "Personal Pronouns (Accusative)"
    # are never collapsed into the same stripped key "personal pronouns".
    def _norm_exact(s: str) -> str:
        """Lowercase + strip trailing level tag only — keep parentheticals."""
        s = s.strip()
        s = re.sub(r'\s+[A-C][12]\s*$', '', s).strip()
        return s.lower()

    exact_to_canon = {
        level: {_norm_exact(t): t for t in topics}
        for level, topics in canonical.items()
    }
    norm_to_canon = {
        level: {_norm(t): t for t in topics}
        for level, topics in canonical.items()
    }

    theory          = {level: {} for level in LEVELS}
    current_level   = None
    current_topic   = None
    body_parts: list[str] = []

    def flush():
        """Save accumulated body text to theory dict, then reset."""
        if current_level and current_topic:
            theory[current_level][current_topic] = "\n".join(body_parts).strip()
        body_parts.clear()

    for para in doc.paragraphs:
        style    = para.style.name
        raw_text = para.text

        # Split on embedded newlines (Bug-fix #5: inline headings in Normal/Web paragraphs)
        parts = raw_text.split('\n')

        for part in parts:
            text = part.strip()
            if not text:
                continue

            # ── Level heading (H1 only) ──────────────────────────────────────
            if style == 'Heading 1' and text in LEVELS:
                flush()
                current_level = text
                current_topic = None
                body_parts.clear()
                continue

            # ── Topic title (any style, for any level) ───────────────────────
            # Also handles: H2, Normal, Normal (Web), even H1 non-level headers
            # (e.g., B1 "Comparative Connectors" mistakenly tagged H1)
            # Try exact match first (preserves parentheticals), then stripped fallback.
            if current_level:
                canon = (exact_to_canon.get(current_level, {}).get(_norm_exact(text))
                         or norm_to_canon.get(current_level, {}).get(_norm(text)))
                if canon:
                    flush()
                    current_topic = canon
                    body_parts.clear()
                    continue

            # ── Body text ────────────────────────────────────────────────────
            if current_level and current_topic:
                body_parts.append(text)

    flush()
    return theory


# ── 4. Parse grammar_exercises.xlsx ──────────────────────────────────────────
SKIP_TOPICS = {
    None, 'Topic', 'German sentence (use ___ for the blank)', 'German_Sentence'
}

def parse_xlsx(xlsx_path: Path, canonical: dict) -> dict:
    """Returns {level: {topic_key: {instruction, exercises:[{sentence,answer,hint?}]}}}."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)

    exercises = {level: {} for level in LEVELS}

    for level in LEVELS:
        if level not in wb.sheetnames:
            continue
        ws = wb[level]
        norm_map = {_norm(t): t for t in canonical.get(level, [])}

        current_key  = None
        last_instr   = "Fill in the blank."

        for row in ws.iter_rows(values_only=True):
            # Ensure we always have exactly 5 columns
            cells = (list(row) + [None] * 5)[:5]
            topic_val, instr_val, sentence_val, answer_val, hint_val = cells

            # Skip header rows
            if topic_val in SKIP_TOPICS:
                continue

            # New topic row (column A has a value)
            if topic_val is not None:
                canon = norm_map.get(_norm(str(topic_val)))
                if canon is None:
                    current_key = None       # unknown topic — skip its rows
                    continue
                current_key = canon
                if instr_val:
                    last_instr = str(instr_val)
                # Initialise entry if not yet seen
                if current_key not in exercises[level]:
                    exercises[level][current_key] = {
                        "instruction": last_instr,
                        "exercises": [],
                    }

            # Continuation row (column A is None) — reuse current_key
            if current_key is None:
                continue

            if not sentence_val or not answer_val:
                continue

            entry: dict = {
                "sentence": str(sentence_val),
                "answer":   str(answer_val),
            }
            if hint_val:
                entry["hint"] = str(hint_val)

            exercises[level][current_key]["exercises"].append(entry)

    return exercises


# ── 5. Write grammarExercises.js ──────────────────────────────────────────────
def write_js(out_path: Path, canonical: dict, theory: dict, exercises: dict):
    lines = [
        "// AUTO-GENERATED — run parse_v3.py to regenerate.",
        "// Sources: grammar_theory.docx  +  grammar_exercises.xlsx",
        "",
        "export const GRAMMAR_EXERCISES = {",
    ]

    for level in LEVELS:
        lines.append(f"  {level}: {{")
        for topic in canonical.get(level, []):
            # Bug-fix #2: ensure_ascii=False keeps ä/ö/ü/etc. as real chars
            key = json.dumps(topic, ensure_ascii=False)

            th   = theory.get(level, {}).get(topic, "")
            # Escape backticks and template-literal ${ sequences
            th_escaped = th.replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")

            ex_data  = exercises.get(level, {}).get(topic, {})
            instr    = ex_data.get("instruction", "Fill in the blank.")
            exs      = ex_data.get("exercises", [])

            lines.append(f"    {key}: {{")
            lines.append(f"      theory: `{th_escaped}`,")
            lines.append(f"      instruction: `{instr}`,")
            lines.append(f"      exercises: [")
            for ex in exs:
                s = json.dumps(ex["sentence"], ensure_ascii=False)
                a = json.dumps(ex["answer"],   ensure_ascii=False)
                if "hint" in ex:
                    h = json.dumps(ex["hint"], ensure_ascii=False)
                    lines.append(f"        {{ sentence: {s}, answer: {a}, hint: {h} }},")
                else:
                    lines.append(f"        {{ sentence: {s}, answer: {a} }},")
            lines.append(f"      ],")
            lines.append(f"    }},")
        lines.append(f"  }},")

    lines.append("};")
    lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  ✓  Written: {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=== parse_v3.py ===\n")

    print("1/3  Loading canonical topics from grammar.js …")
    canonical = load_canonical_topics(GRAMMAR_JS)
    for level, topics in canonical.items():
        print(f"     {level}: {len(topics)} topics")

    print("\n2/3  Parsing grammar_theory.docx …")
    theory = parse_docx(DOCX_PATH, canonical)
    for level in LEVELS:
        filled = sum(1 for v in theory[level].values() if v)
        total  = len(canonical.get(level, []))
        status = [
            f"  {'✓' if v else '✗ EMPTY'} {t}"
            for t, v in [
                (topic, theory[level].get(topic, ""))
                for topic in canonical.get(level, [])
            ]
        ]
        print(f"     {level}: {filled}/{total} topics have theory")
        for s in status:
            print(f"         {s}")

    print("\n3/3  Parsing grammar_exercises.xlsx …")
    exercises = parse_xlsx(XLSX_PATH, canonical)
    for level in LEVELS:
        topics_with_ex = len(exercises[level])
        total_ex = sum(len(v["exercises"]) for v in exercises[level].values())
        unmatched = [
            t for t in exercises[level]
            if t not in canonical.get(level, [])
        ]
        print(f"     {level}: {topics_with_ex} topics, {total_ex} exercises")
        if unmatched:
            print(f"     WARNING — unmatched topic keys: {unmatched}")

    print("\n4/4  Writing grammarExercises.js …")
    write_js(OUT_JS, canonical, theory, exercises)

    print("\nDone ✓")
